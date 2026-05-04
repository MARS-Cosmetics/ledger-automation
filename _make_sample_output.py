"""Produces sample_output.xlsx from the test fixtures so you can preview the
output format without needing streamlit. Not part of the app — utility only."""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reconcile import reconcile
from test_reconcile import build_fixtures


def build_excel(res, path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        stats_df = pd.DataFrame(
            [
                ["Mars rows in Recon period", res.stats["mars_recon_rows"]],
                ["Brand rows in Recon period", res.stats["brand_recon_rows"]],
                ["Mars rows matched", res.stats["mars_matched_rows"]],
                ["Mars rows not booked by Brand", res.stats["mars_unmatched_rows"]],
                ["Brand rows not booked by Mars", res.stats["brand_unmatched_rows"]],
            ],
            columns=["Metric", "Value"],
        )
        stats_df.to_excel(writer, sheet_name="Recon Summary", index=False, startrow=0)
        start_row = len(stats_df) + 2
        res.summary.to_excel(writer, sheet_name="Recon Summary", index=False, startrow=start_row)
        res.mars.to_excel(writer, sheet_name="Mars", index=False)
        res.brand.to_excel(writer, sheet_name="Brand", index=False)

        wb = writer.book
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        not_booked_fill = PatternFill("solid", fgColor="FCE4D6")
        money_fmt = "#,##0.00;[Red]-#,##0.00"

        for sn in ("Recon Summary", "Mars", "Brand"):
            ws = wb[sn]
            for ci, cells in enumerate(ws.columns, start=1):
                m = max((len(str(c.value or "")) for c in cells), default=10)
                ws.column_dimensions[get_column_letter(ci)].width = min(max(12, m + 2), 40)

        ws = wb["Recon Summary"]
        for c in ws[1]:
            c.font = bold; c.fill = header_fill
        for c in ws[start_row + 1]:
            c.font = bold; c.fill = header_fill
        for r in range(start_row + 2, start_row + 2 + len(res.summary)):
            for col in (2, 3, 4):
                ws.cell(row=r, column=col).number_format = money_fmt
            if ws.cell(row=r, column=1).value == "Total":
                for col in range(1, 5):
                    ws.cell(row=r, column=col).font = bold
                    ws.cell(row=r, column=col).fill = total_fill

        for sn, df in (("Mars", res.mars), ("Brand", res.brand)):
            ws = wb[sn]
            for c in ws[1]:
                c.font = bold; c.fill = header_fill
            ws.freeze_panes = "A2"
            cols = list(df.columns)
            money_idx = []
            for cn in cols:
                lc = cn.lower()
                if cn in ("Amount_Brand", "Amount_Mars", "Difference", "Diff") or "amount" in lc or lc in ("debit", "credit"):
                    money_idx.append(cols.index(cn) + 1)
            try:
                rem_idx = cols.index("Remarks") + 1
            except ValueError:
                rem_idx = None
            for r in range(2, 2 + len(df)):
                for c in money_idx:
                    ws.cell(row=r, column=c).number_format = money_fmt
                if rem_idx and "Not Booked" in str(ws.cell(row=r, column=rem_idx).value or ""):
                    for c in range(1, len(cols) + 1):
                        ws.cell(row=r, column=c).fill = not_booked_fill


if __name__ == "__main__":
    mars, brand = build_fixtures()
    cols_m = {"vch_no": "Vch No", "net_amount": "Net Amount", "period": "Period", "category": "Category"}
    cols_b = {"reference": "Reference", "net_amount": "Net Amount", "period": "Period", "category": "Category"}
    res = reconcile(mars, brand, cols_m, cols_b)
    build_excel(res, "sample_output.xlsx")
    print("Wrote sample_output.xlsx")
    print(res.summary.to_string(index=False))
