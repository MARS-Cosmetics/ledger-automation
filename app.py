"""Streamlit UI for the Mars vs Brand reconciliation tool.

Run: streamlit run app.py    (or use ./run.sh)
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from column_detect import (
    BRAND_FIELD_SYNONYMS,
    MARS_FIELD_SYNONYMS,
    REQUIRED_BRAND,
    REQUIRED_MARS,
    detect_columns,
    missing_required,
)
from reconcile import ReconResult, reconcile


st.set_page_config(page_title="Mars Recon Tool", layout="wide")
st.title("Mars vs Brand — Ledger Reconciliation")
st.caption(
    "Drop both ledgers below. The tool auto-detects columns — confirm the mapping "
    "before running so a wrong guess doesn't silently break the recon."
)

# ----- File upload -----
c1, c2 = st.columns(2)
with c1:
    mars_file = st.file_uploader(
        "Mars ledger (.xlsx / .xls)", type=["xlsx", "xls"], key="mars"
    )
with c2:
    brand_file = st.file_uploader(
        "Brand ledger (.xlsx / .xls)", type=["xlsx", "xls"], key="brand"
    )


@st.cache_data(show_spinner=False)
def _read_excel(file_bytes: bytes, sheet=None) -> dict:
    """Return all sheets as a dict {sheet_name: df}."""
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)


def _pick_sheet(label: str, sheets: dict, key: str) -> pd.DataFrame:
    names = list(sheets.keys())
    if len(names) == 1:
        return sheets[names[0]]
    pick = st.selectbox(f"{label} — sheet", names, key=key)
    return sheets[pick]


def _column_mapper(label: str, df: pd.DataFrame, synonyms: dict, required: list, prefix: str) -> dict:
    """Render selectboxes for each logical field; return {field: column_or_None}."""
    detected = detect_columns(df, synonyms)
    st.markdown(f"**{label}**")
    cols_per_row = 4
    fields = list(synonyms.keys())
    mapping: dict = {}
    options = ["(none)"] + list(df.columns)

    for i in range(0, len(fields), cols_per_row):
        row = st.columns(cols_per_row)
        for j, field in enumerate(fields[i : i + cols_per_row]):
            with row[j]:
                default = detected.get(field) or "(none)"
                idx = options.index(default) if default in options else 0
                req_marker = " *" if field in required else ""
                pick = st.selectbox(
                    f"{field}{req_marker}", options, index=idx, key=f"{prefix}_{field}"
                )
                if pick != "(none)":
                    mapping[field] = pick
    return mapping


def _build_excel_bytes(res: ReconResult) -> bytes:
    """Write summary + mars + brand sheets with light formatting."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ---- Summary sheet (with stats block on top) ----
        summary_sheet = "Recon Summary"
        stats_df = pd.DataFrame(
            [
                ["Mars rows in Recon period", res.stats["mars_recon_rows"]],
                ["Brand rows in Recon period", res.stats["brand_recon_rows"]],
                ["Mars rows matched", res.stats["mars_matched_rows"]],
                ["Mars rows not booked by Brand", res.stats["mars_unmatched_rows"]],
                ["Brand rows not booked by Mars", res.stats["brand_unmatched_rows"]],
                ["Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ],
            columns=["Metric", "Value"],
        )
        stats_df.to_excel(writer, sheet_name=summary_sheet, index=False, startrow=0)
        # blank row, then summary table
        start_row = len(stats_df) + 2
        res.summary.to_excel(writer, sheet_name=summary_sheet, index=False, startrow=start_row)

        res.mars.to_excel(writer, sheet_name="Mars", index=False)
        res.brand.to_excel(writer, sheet_name="Brand", index=False)

        # ---- Formatting ----
        wb = writer.book
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        not_booked_fill = PatternFill("solid", fgColor="FCE4D6")
        money_fmt = "#,##0.00;[Red]-#,##0.00"

        for sheet_name in (summary_sheet, "Mars", "Brand"):
            ws = wb[sheet_name]
            # auto-width
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 40)

        # Summary sheet styling
        ws = wb[summary_sheet]
        # bold the two header rows (stats header at row 1, summary header at start_row+1)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
        for cell in ws[start_row + 1]:
            cell.font = bold
            cell.fill = header_fill
        # money format for summary amount columns + Total row highlight
        for r in range(start_row + 2, start_row + 2 + len(res.summary)):
            for c in (2, 3, 4):  # Amount_Mars, Amount_Brand, Difference
                ws.cell(row=r, column=c).number_format = money_fmt
            if ws.cell(row=r, column=1).value == "Total":
                for c in range(1, 5):
                    ws.cell(row=r, column=c).font = bold
                    ws.cell(row=r, column=c).fill = total_fill

        # Mars / Brand styling: header row, money format on amount cols, highlight unmatched rows
        for sheet_name, df in (("Mars", res.mars), ("Brand", res.brand)):
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = bold
                cell.fill = header_fill
            ws.freeze_panes = "A2"

            cols = list(df.columns)
            money_col_names = []
            for cname in cols:
                lc = cname.lower()
                if cname in ("Amount_Brand", "Amount_Mars", "Difference", "Diff"):
                    money_col_names.append(cname)
                elif "amount" in lc or lc in ("debit", "credit"):
                    money_col_names.append(cname)
            money_col_idx = [cols.index(n) + 1 for n in money_col_names]
            try:
                remarks_idx = cols.index("Remarks") + 1
            except ValueError:
                remarks_idx = None

            for r in range(2, 2 + len(df)):
                for c in money_col_idx:
                    ws.cell(row=r, column=c).number_format = money_fmt
                if remarks_idx:
                    rem = ws.cell(row=r, column=remarks_idx).value or ""
                    if "Not Booked" in str(rem):
                        for c in range(1, len(cols) + 1):
                            ws.cell(row=r, column=c).fill = not_booked_fill

    buf.seek(0)
    return buf.read()


# ----- Main flow -----
if mars_file and brand_file:
    try:
        mars_sheets = _read_excel(mars_file.getvalue())
        brand_sheets = _read_excel(brand_file.getvalue())
    except Exception as e:
        st.error(f"Could not read one of the files: {e}")
        st.stop()

    st.divider()
    st.subheader("1. Pick the right sheet")
    sc1, sc2 = st.columns(2)
    with sc1:
        mars_df = _pick_sheet("Mars", mars_sheets, "mars_sheet")
        st.caption(f"{len(mars_df)} rows · columns: {', '.join(mars_df.columns.astype(str))}")
    with sc2:
        brand_df = _pick_sheet("Brand", brand_sheets, "brand_sheet")
        st.caption(f"{len(brand_df)} rows · columns: {', '.join(brand_df.columns.astype(str))}")

    st.divider()
    st.subheader("2. Confirm column mapping")
    st.caption("Required fields are marked with `*`. Override any wrong guess before reconciling.")

    mars_cols = _column_mapper("Mars ledger", mars_df, MARS_FIELD_SYNONYMS, REQUIRED_MARS, "mars")
    st.markdown("&nbsp;", unsafe_allow_html=True)
    brand_cols = _column_mapper("Brand ledger", brand_df, BRAND_FIELD_SYNONYMS, REQUIRED_BRAND, "brand")

    miss_m = missing_required(mars_cols, REQUIRED_MARS)
    miss_b = missing_required(brand_cols, REQUIRED_BRAND)

    st.divider()
    st.subheader("3. Run reconciliation")

    if miss_m or miss_b:
        msg = []
        if miss_m:
            msg.append(f"Mars missing: {', '.join(miss_m)}")
        if miss_b:
            msg.append(f"Brand missing: {', '.join(miss_b)}")
        st.warning("Map all required (*) fields before running. " + " · ".join(msg))
    else:
        if st.button("Reconcile", type="primary"):
            try:
                with st.spinner("Reconciling…"):
                    res = reconcile(mars_df, brand_df, mars_cols, brand_cols)
            except Exception as e:
                st.error(f"Reconciliation failed: {e}")
                st.stop()

            st.success("Done.")

            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Mars (Recon) rows", res.stats["mars_recon_rows"])
            mc2.metric("Brand (Recon) rows", res.stats["brand_recon_rows"])
            mc3.metric("Mars not booked by Brand", res.stats["mars_unmatched_rows"])
            mc4.metric("Brand not booked by Mars", res.stats["brand_unmatched_rows"])

            st.markdown("**Recon Summary (preview)**")
            st.dataframe(res.summary, use_container_width=True)

            with st.expander("Mars sheet preview"):
                st.dataframe(res.mars.head(50), use_container_width=True)
            with st.expander("Brand sheet preview"):
                st.dataframe(res.brand.head(50), use_container_width=True)

            xlsx = _build_excel_bytes(res)
            fname = f"recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                "Download recon workbook",
                data=xlsx,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
else:
    st.info("Upload both files to begin.")
